"""
PowerPulse — Monthly Energy Report Generator

Usage:
    python reports/build_monthly_report.py 2008-03

If no month is given, defaults to the most recent complete month in the data.
Output: reports/monthly_energy_report_<month>.xlsx
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# --- Configuration ---------------------------------------------------------
DATA_PATH      = Path("data/processed/power_readings_clean.parquet")
OUTPUT_DIR     = Path("reports")
PEAK_RATE_EUR     = 0.20   # default peak rate; user can change in the Summary sheet
OFF_PEAK_RATE_EUR = 0.12
PEAK_HOURS = range(6, 22)   # 06:00 inclusive to 22:00 exclusive
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
KPI_LABEL_FONT = Font(name="Calibri", bold=True, size=11)
KPI_VALUE_FONT = Font(name="Calibri", size=11)

def load_month(month_str: str | None) -> tuple[str, pd.DataFrame]:
    """Load cleaned data; return (month_str, df_for_that_month)."""
    df = pd.read_parquet(DATA_PATH)

    if month_str is None:
        # Default: the most recent month that has at least 25 days of data
        completeness = df.resample("ME")["global_active_power"].count() / (60 * 24)
        complete_months = completeness[completeness >= 25].index
        if len(complete_months) == 0:
            raise SystemExit("No complete months found in data.")
        month_str = complete_months.max().strftime("%Y-%m")

    month_df = df.loc[month_str].copy()
    if month_df.empty:
        raise SystemExit(f"No data for month {month_str}.")
    return month_str, month_df

def daily_table(month_df: pd.DataFrame) -> pd.DataFrame:
    """One row per day with kWh totals and peak/off-peak split."""
    month_df = month_df.copy()
    month_df["kwh"]      = month_df["total_wh"] / 1000.0
    month_df["is_peak"]  = month_df.index.hour.isin(PEAK_HOURS)
    month_df["peak_kwh"]    = month_df["kwh"].where(month_df["is_peak"], 0)
    month_df["offpeak_kwh"] = month_df["kwh"].where(~month_df["is_peak"], 0)

    daily = month_df.resample("D").agg(
        total_kwh=("kwh", "sum"),
        peak_kwh=("peak_kwh", "sum"),
        offpeak_kwh=("offpeak_kwh", "sum"),
    ).round(3)
    daily.index = daily.index.date  # cleaner display in Excel
    return daily

def write_summary(ws, month_str, daily, n_days):
    ws.title = "Summary"

    # Title
    ws["A1"] = f"Monthly Energy Report — {month_str}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16)
    ws.merge_cells("A1:D1")

    # Tariff parameters (editable — these are the cells stakeholders can change)
    ws["A3"] = "Tariff parameters"
    ws["A3"].font = KPI_LABEL_FONT
    ws["A4"] = "Peak rate (€/kWh)"     ; ws["B4"] = PEAK_RATE_EUR
    ws["A5"] = "Off-peak rate (€/kWh)" ; ws["B5"] = OFF_PEAK_RATE_EUR
    for cell in (ws["B4"], ws["B5"]):
        cell.number_format = "0.00"
        cell.fill = PatternFill("solid", start_color="FFF2CC")  # yellow = "editable"

    # KPIs — written as FORMULAS referencing the Daily sheet, so they recalc
    # automatically when the daily table is edited.
    last_row = 1 + n_days + 1   # daily header row + n data rows + 1 totals row
    daily_range_total = f"'Daily breakdown'!B2:B{1+n_days}"
    daily_range_peak  = f"'Daily breakdown'!C2:C{1+n_days}"
    daily_range_offpk = f"'Daily breakdown'!D2:D{1+n_days}"
    daily_range_date  = f"'Daily breakdown'!A2:A{1+n_days}"

    ws["A7"] = "Key metrics" ; ws["A7"].font = KPI_LABEL_FONT
    ws["A8"]  = "Total energy (kWh)"      ; ws["B8"]  = f"=SUM({daily_range_total})"
    ws["A9"]  = "Peak-hour energy (kWh)"  ; ws["B9"]  = f"=SUM({daily_range_peak})"
    ws["A10"] = "Off-peak energy (kWh)"   ; ws["B10"] = f"=SUM({daily_range_offpk})"
    ws["A11"] = "Average daily energy (kWh)" ; ws["B11"] = f"=AVERAGE({daily_range_total})"
    ws["A12"] = "Peak day (kWh)"          ; ws["B12"] = f"=MAX({daily_range_total})"
    ws["A13"] = "Peak day date"           ; ws["B13"] = (
        f"=INDEX({daily_range_date},MATCH(MAX({daily_range_total}),{daily_range_total},0))"
    )
    ws["A14"] = "Total cost (€)"          ; ws["B14"] = (
        f"=SUM({daily_range_peak})*$B$4 + SUM({daily_range_offpk})*$B$5"
    )

    # Formatting
    for r in range(8, 15):
        ws.cell(row=r, column=1).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2).font = KPI_VALUE_FONT
        if r != 13:
            ws.cell(row=r, column=2).number_format = "#,##0.00"
    ws["B14"].number_format = '"€"#,##0.00'

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

def write_daily(ws, daily):
    ws.title = "Daily breakdown"
    headers = ["Date", "Total kWh", "Peak kWh", "Off-peak kWh", "Daily cost (€)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, (day, row) in enumerate(daily.iterrows(), start=2):
        ws.cell(row=i, column=1, value=day)
        ws.cell(row=i, column=2, value=float(row["total_kwh"]))
        ws.cell(row=i, column=3, value=float(row["peak_kwh"]))
        ws.cell(row=i, column=4, value=float(row["offpeak_kwh"]))
        # Cost = peak*peak_rate + offpeak*offpeak_rate (rates from Summary sheet)
        ws.cell(row=i, column=5,
                value=f"=C{i}*Summary!$B$4 + D{i}*Summary!$B$5")

    # Totals row
    total_row = len(daily) + 2
    ws.cell(row=total_row, column=1, value="Total").font = KPI_LABEL_FONT
    for col in range(2, 6):
        L = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({L}2:{L}{total_row-1})").font = KPI_LABEL_FONT

    # Formatting
    for r in range(2, total_row + 1):
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).number_format = "#,##0.00"
        ws.cell(row=r, column=5).number_format = '"€"#,##0.00'

    for col_letter, width in [("A",12), ("B",12), ("C",12), ("D",14), ("E",16)]:
        ws.column_dimensions[col_letter].width = width

def write_load_curve(ws, month_df):
    ws.title = "Hourly load curve"
    hourly = (month_df["global_active_power"]
              .groupby(month_df.index.hour).mean().round(3))

    ws.append(["Hour of day", "Average active power (kW)"])
    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, (hour, kw) in enumerate(hourly.items(), start=2):
        ws.cell(row=i, column=1, value=int(hour))
        ws.cell(row=i, column=2, value=float(kw))
        ws.cell(row=i, column=2).number_format = "0.000"

    # Embed a line chart
    chart = LineChart()
    chart.title = "Average power by hour of day"
    chart.y_axis.title = "Average active power (kW)"
    chart.x_axis.title = "Hour of day"
    chart.height = 10
    chart.width = 18
    data = Reference(ws, min_col=2, min_row=1, max_row=25, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=25)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26

def main():
    month_str = sys.argv[1] if len(sys.argv) > 1 else None
    month_str, month_df = load_month(month_str)
    daily = daily_table(month_df)

    wb = Workbook()
    write_summary(wb.active, month_str, daily, n_days=len(daily))
    write_daily(wb.create_sheet(), daily)
    write_load_curve(wb.create_sheet(), month_df)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"monthly_energy_report_{month_str}.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")

if __name__ == "__main__":
    main()